from flask import Flask, request, render_template, redirect
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = 'test_request.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Requests"
        headers = [
            "Date", "Customer Name", "Customer Address", "Contact Number", "Email",
            "UUT Name", "UUT Quantity", "Dimension", "Weight", "UUT Sr. No.", "Repeat Test", "Ref No.",
            "Test Name", "Test Specification", "Test Standard",
            "Special Requirement"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/form')
def form():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    data = request.form
    
    row = [
        data.get("date"),
        data.get("c_name"),
        data.get("c_address"),
        data.get("c_number"),
        data.get("email"),

        data.get("uut_name"),
        data.get("uut_quantity"),
        data.get("dimension"),
        data.get("weight"),
        data.get("uut_sr_no"),
        data.get("repeat_test"),
        data.get("ref_no"),
        
        data.get("test_name"),
        data.get("test_specification"),
        data.get("test_standard"),
        data.get("special_requirement")
    ]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(row)
    wb.save(EXCEL_FILE)

    return "Success",200

if __name__ == '__main__':
    init_excel()
    app.run(debug=True)
